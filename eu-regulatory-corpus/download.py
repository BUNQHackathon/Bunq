#!/usr/bin/env python3
"""
download.py — EU Regulatory Corpus downloader.

Strategy for EUR-Lex: use the EU Publications Office SPARQL endpoint
(https://publications.europa.eu/webapi/rdf/sparql) to resolve CELEX IDs to
cellar manifestation URIs, then download PDFs via
https://publications.europa.eu/resource/cellar/{id}/DOC_1.

Strategy for EBA: scrape listing pages at eba.europa.eu/publications-and-media/publications
filtered by document type (Guidelines=250, Opinions=252, Recommendations=255),
paginate through all result pages, and download the primary PDF for each item.

Usage:
    python download.py                 # download all sources
    python download.py --source eurlex # EUR-Lex only
    python download.py --source eba    # EBA Guidelines/Recommendations/Opinions only
"""

import argparse
import csv
import hashlib
import io
import json
import logging
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urljoin, urlparse

import httpx
import yaml
from bs4 import BeautifulSoup
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
DOCS_PRIMARY = BASE_DIR / "docs" / "primary"
DOCS_EBA = BASE_DIR / "docs" / "guidance" / "eba"
LOGS_DIR = BASE_DIR / "logs"
INDEX_CSV = BASE_DIR / "index.csv"
SOURCES_YAML = BASE_DIR / "sources.yaml"

SPARQL_ENDPOINT = "https://publications.europa.eu/webapi/rdf/sparql"
CELLAR_BASE = "https://publications.europa.eu/resource/cellar"

EBA_BASE = "https://www.eba.europa.eu"
# document_type IDs on the EBA publications listing
EBA_CATEGORIES = {
    "guideline": 250,
    "recommendation": 255,
    "opinion": 252,
}

ECB_BASE = "https://www.bankingsupervision.europa.eu"
DOCS_ECB = BASE_DIR / "docs" / "guidance" / "ecb"

# foedb API constants for ECB Banking Supervision publications database
ECB_FOEDB_HOST = "/foedb/dbs/foedb"
ECB_FOEDB_DB = "publications.en"
ECB_FOEDB_HEADER = [
    "id", "pub_timestamp", "year", "issue_number", "type",
    "boardmember", "Authors", "documentTypes", "publicationProperties",
    "childrenPublication", "relatedPublications",
]
# foedb type IDs for categories we care about
ECB_TYPE_GUIDE = 305        # Supervisory Guides
ECB_TYPE_PRIORITIES = 186   # Supervisory priorities
# All "publication" types (non-speeches, non-interviews, non-MEP-letters)
# Exclude: 47=press release, 48=working paper, 49=interview, 91=opinion,
#          94=letter to MEP, 192=news, 215=hearing, 227=blog, 244=FAQ
ECB_PUB_EXCLUDE_TYPES = {47, 48, 49, 91, 94, 192, 215, 227, 244}
# Minimum year to crawl (start of SSM)
ECB_MIN_YEAR = 2014

CSV_COLUMNS = [
    "id", "source", "short_name", "title", "doc_type",
    "url", "published_date", "local_path", "sha256",
    "downloaded_at", "status",
]

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/131.0.0.0 Safari/537.36"
)
MIN_INTERVAL = 1.5  # seconds between requests to the same host

console = Console()

# ---------------------------------------------------------------------------
# EBA module logger
# ---------------------------------------------------------------------------
LOGS_DIR.mkdir(parents=True, exist_ok=True)
_eba_logger = logging.getLogger("eba_crawler")
_eba_logger.setLevel(logging.INFO)
if not _eba_logger.handlers:
    _fh = logging.FileHandler(LOGS_DIR / "eba.log", encoding="utf-8")
    _fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _eba_logger.addHandler(_fh)

# ---------------------------------------------------------------------------
# ECB module logger
# ---------------------------------------------------------------------------
_ecb_logger = logging.getLogger("ecb_crawler")
_ecb_logger.setLevel(logging.INFO)
if not _ecb_logger.handlers:
    _fh_ecb = logging.FileHandler(LOGS_DIR / "ecb.log", encoding="utf-8")
    _fh_ecb.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _ecb_logger.addHandler(_fh_ecb)

# ---------------------------------------------------------------------------
# FATF constants
# ---------------------------------------------------------------------------
FATF_BASE = "https://www.fatf-gafi.org"
DOCS_FATF = BASE_DIR / "docs" / "guidance" / "fatf"

# AEM faceted-search API endpoint (server-rendered, behind Cloudflare — Playwright required)
FATF_SEARCH_API = (
    "/content/fatf-gafi/en/search-page"
    "/jcr:content/root/container/faceted_search/results.facets.json"
)
# doc-type facet values to crawl (from the checkboxes on /en/search-page.html)
FATF_DOC_TYPE_FILTERS = {
    "guidance":         "fatf-gafi-faft-doc types:tag-Guidance",
    "recommendations":  "fatf-gafi-faft-doc types:tag-Recommendations",
    "methodology":      "fatf-gafi-faft-doc types:tag-Risk Based Approach",
}
# Categories that are out of scope regardless of doc-type tag
FATF_SKIP_CATEGORIES = {
    "Mutualevaluations",
    "High-risk-and-other-monitored-jurisdictions",
}

# ---------------------------------------------------------------------------
# FATF module logger
# ---------------------------------------------------------------------------
_fatf_logger = logging.getLogger("fatf_crawler")
_fatf_logger.setLevel(logging.INFO)
if not _fatf_logger.handlers:
    _fh_fatf = logging.FileHandler(LOGS_DIR / "fatf.log", encoding="utf-8")
    _fh_fatf.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _fatf_logger.addHandler(_fh_fatf)

# ---------------------------------------------------------------------------
# ESMA constants
# ---------------------------------------------------------------------------
ESMA_BASE = "https://www.esma.europa.eu"
DOCS_ESMA = BASE_DIR / "docs" / "guidance" / "esma"

ESMA_GUIDELINES_TRACKER_URL = (
    f"{ESMA_BASE}/sites/default/files/library/guidelines_tracker.xlsx"
)
ESMA_QA_BUNDLES_URL = f"{ESMA_BASE}/publications-and-data/questions-answers"

# Two-letter EU language suffixes to strip when finding the English PDF
_ESMA_LANG_SUFFIXES = {
    "BG", "CS", "DA", "DE", "EL", "ES", "ET", "FI", "FR", "GA",
    "HR", "HU", "IT", "LT", "LV", "MT", "NL", "PL", "PT", "RO",
    "SK", "SL", "SV",
}

# ESMA reference pattern: ESMAnn-nnnnn-nnn  or  ESMAnn_nnnnn_nnn  etc.
_ESMA_REF_RE = re.compile(r"(ESMA[\d]+[-_][\d]+(?:[-_][\d]+)?)", re.IGNORECASE)

# ---------------------------------------------------------------------------
# ESMA module logger
# ---------------------------------------------------------------------------
_esma_logger = logging.getLogger("esma_crawler")
_esma_logger.setLevel(logging.INFO)
if not _esma_logger.handlers:
    _fh_esma = logging.FileHandler(LOGS_DIR / "esma.log", encoding="utf-8")
    _fh_esma.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _esma_logger.addHandler(_fh_esma)

# ---------------------------------------------------------------------------
# Irish Statute Book (national/ireland/primary)
# ---------------------------------------------------------------------------
ISB_BASE = "https://www.irishstatutebook.ie"
LRC_REVISED_BASE = "https://revisedacts.lawreform.ie"
DOCS_IE_PRIMARY = BASE_DIR / "docs" / "national" / "ireland" / "primary"
DOCS_IE_CBI = BASE_DIR / "docs" / "national" / "ireland" / "guidance" / "cbi"
DOCS_IE_DPC = BASE_DIR / "docs" / "national" / "ireland" / "guidance" / "dpc"

_isb_logger = logging.getLogger("isb_crawler")
_isb_logger.setLevel(logging.INFO)
if not _isb_logger.handlers:
    _fh_isb = logging.FileHandler(LOGS_DIR / "irishstatutebook.log", encoding="utf-8")
    _fh_isb.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _isb_logger.addHandler(_fh_isb)

# ---------------------------------------------------------------------------
# Central Bank of Ireland (national/ireland/guidance/cbi)
# ---------------------------------------------------------------------------
CBI_BASE = "https://www.centralbank.ie"

_cbi_logger = logging.getLogger("cbi_crawler")
_cbi_logger.setLevel(logging.INFO)
if not _cbi_logger.handlers:
    _fh_cbi = logging.FileHandler(LOGS_DIR / "cbi.log", encoding="utf-8")
    _fh_cbi.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    _cbi_logger.addHandler(_fh_cbi)

# ---------------------------------------------------------------------------
# Rate limiter
# ---------------------------------------------------------------------------
_last_request: dict[str, float] = {}


def _throttle(host: str) -> None:
    now = time.monotonic()
    wait = MIN_INTERVAL - (now - _last_request.get(host, 0))
    if wait > 0:
        time.sleep(wait)
    _last_request[host] = time.monotonic()


# ---------------------------------------------------------------------------
# HTTP helpers
# ---------------------------------------------------------------------------
def _get(client: httpx.Client, url: str, **kwargs) -> httpx.Response:
    host = httpx.URL(url).host
    _throttle(host)
    return client.get(url, follow_redirects=True, timeout=120, **kwargs)


# ---------------------------------------------------------------------------
# Index CSV helpers
# ---------------------------------------------------------------------------
def load_index() -> dict[str, dict]:
    """Return {celex_id: row_dict} from index.csv."""
    rows: dict[str, dict] = {}
    if INDEX_CSV.exists():
        with INDEX_CSV.open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                rows[row["id"]] = row
    return rows


def save_index(rows: dict[str, dict]) -> None:
    with INDEX_CSV.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        for row in rows.values():
            writer.writerow(row)


def upsert_index_row(row: dict) -> None:
    """Atomically upsert a single row into index.csv, keyed on row['id'].
    Loads existing rows, replaces any with matching id, appends new, writes via tmp+rename."""
    rows = []
    if INDEX_CSV.exists():
        with open(INDEX_CSV, "r", newline="", encoding="utf-8") as f:
            rows = [r for r in csv.DictReader(f) if r.get("id") != row["id"]]
    rows.append({k: row.get(k, "") for k in CSV_COLUMNS})
    tmp = INDEX_CSV.with_suffix(".csv.tmp")
    with open(tmp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        w.writeheader()
        w.writerows(rows)
    tmp.replace(INDEX_CSV)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def doc_type_from_celex(celex: str) -> str:
    """Derive doc_type from CELEX type character (R=regulation, L=directive)."""
    m = re.search(r"\d{4}([A-Z])", celex)
    if m:
        t = m.group(1)
        if t == "R":
            return "regulation"
        if t == "L":
            return "directive"
    return "unknown"


# ---------------------------------------------------------------------------
# SPARQL helpers
# ---------------------------------------------------------------------------
def sparql_query(client: httpx.Client, query: str) -> list[dict]:
    """Run a SPARQL SELECT query and return the bindings list."""
    resp = _get(
        client,
        SPARQL_ENDPOINT,
        params={"query": query, "format": "application/json"},
    )
    resp.raise_for_status()
    return resp.json()["results"]["bindings"]


def find_latest_consolidated(client: httpx.Client, celex: str) -> tuple[str, str] | tuple[None, None]:
    """
    Return (consolidated_celex_id, cellar_work_uri) for the most recent
    consolidated version of the given original CELEX ID, or (None, None).

    Consolidated CELEX IDs start with '0' followed by the digits/letters after
    the leading '3' of the original, e.g. 32016R0679 → 02016R0679-YYYYMMDD.
    """
    fragment = celex[1:]  # strip leading '3'
    query = f"""
PREFIX cdm: <http://publications.europa.eu/ontology/cdm#>
SELECT DISTINCT ?work ?id
WHERE {{
  ?work cdm:work_id_document ?id .
  FILTER(CONTAINS(STR(?id), "0{fragment}"))
}}
ORDER BY DESC(STR(?id))
LIMIT 1
"""
    rows = sparql_query(client, query)
    if not rows:
        return None, None
    celex_val = rows[0]["id"]["value"]
    # Strip "celex:" prefix if present
    celex_val = re.sub(r"^celex:", "", celex_val, flags=re.IGNORECASE)
    work_uri = rows[0]["work"]["value"]
    return celex_val, work_uri


def find_original_work(client: httpx.Client, celex: str) -> tuple[str, str] | tuple[None, None]:
    """Return (celex_id, cellar_work_uri) for the original act."""
    query = f"""
PREFIX cdm: <http://publications.europa.eu/ontology/cdm#>
SELECT DISTINCT ?work ?id
WHERE {{
  ?work cdm:work_id_document ?id .
  FILTER(STR(?id) = "celex:{celex}")
}}
LIMIT 1
"""
    rows = sparql_query(client, query)
    if not rows:
        return None, None
    celex_val = re.sub(r"^celex:", "", rows[0]["id"]["value"], flags=re.IGNORECASE)
    work_uri = rows[0]["work"]["value"]
    return celex_val, work_uri


def find_pdf_manifestation(client: httpx.Client, work_uri: str) -> str | None:
    """
    Return the HTTPS URI of the best English PDF manifestation for the given
    work, or None if none found. Prefers pdfa1a/pdfa2a over pdf.
    """
    query = f"""
PREFIX cdm: <http://publications.europa.eu/ontology/cdm#>
SELECT DISTINCT ?manif ?format
WHERE {{
  ?expr cdm:expression_belongs_to_work <{work_uri}> ;
        cdm:expression_uses_language <http://publications.europa.eu/resource/authority/language/ENG> .
  ?manif cdm:manifestation_manifests_expression ?expr ;
         cdm:manifestation_type ?format .
}}
"""
    rows = sparql_query(client, query)
    if not rows:
        return None

    # Priority: pdfa1a > pdfa2a > pdf > anything
    priority = {"pdfa1a": 0, "pdfa2a": 1, "pdf": 2}
    ranked = sorted(
        rows,
        key=lambda r: priority.get(r["format"]["value"], 99),
    )
    best = ranked[0]
    manif_uri = best["manif"]["value"].replace("http://", "https://")
    return manif_uri  # caller will append /DOC_N or use content negotiation


def download_pdf_from_manifestation(
    client: httpx.Client, manif_uri: str
) -> bytes:
    """
    Download the PDF from a cellar manifestation URI.
    Tries /DOC_1, then /DOC_2, then content negotiation (Accept: application/pdf).
    Raises on complete failure.
    """
    for doc in ("DOC_1", "DOC_2"):
        url = f"{manif_uri}/{doc}"
        _throttle(httpx.URL(url).host)
        r = client.get(url, follow_redirects=True, timeout=120)
        if r.status_code == 200 and r.content.startswith(b"%PDF"):
            return r.content
        # 404 or non-PDF → try next

    # Content negotiation fallback
    _throttle(httpx.URL(manif_uri).host)
    r = client.get(
        manif_uri,
        follow_redirects=True,
        timeout=120,
        headers={"Accept": "application/pdf"},
    )
    r.raise_for_status()
    if not r.content.startswith(b"%PDF"):
        raise ValueError(
            f"Content-negotiated response is not a PDF (first bytes: {r.content[:20]!r})"
        )
    return r.content


# ---------------------------------------------------------------------------
# EUR-Lex downloader
# ---------------------------------------------------------------------------
def download_eurlex(sources: list[dict], index: dict[str, dict]) -> tuple[int, int, int]:
    new_count = unchanged = failed = 0
    DOCS_PRIMARY.mkdir(parents=True, exist_ok=True)

    headers = {"User-Agent": USER_AGENT}

    with httpx.Client(headers=headers) as client:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("Downloading EUR-Lex...", total=len(sources))

            for entry in sources:
                celex = entry["celex"]
                short_name = entry["short_name"]
                progress.update(task, description=f"[cyan]{celex}[/cyan] ({short_name})")

                try:
                    # --- Step 1: resolve work URI (consolidated preferred) ---
                    # Try consolidated first; if it has no English PDF, fall back to original.
                    cons_celex, cons_work_uri = find_latest_consolidated(client, celex)
                    manif_uri: str | None = None
                    actual_celex: str = celex

                    if cons_celex and cons_work_uri:
                        manif_uri = find_pdf_manifestation(client, cons_work_uri)
                        if manif_uri:
                            actual_celex = cons_celex
                            console.log(f"  [green]consolidated[/green] {celex} → {cons_celex}")
                        else:
                            console.log(
                                f"  [yellow]consolidated exists but no PDF[/yellow] "
                                f"for {cons_celex}, falling back to original"
                            )

                    if not manif_uri:
                        # No consolidated PDF — use original act
                        orig_celex, orig_work_uri = find_original_work(client, celex)
                        if not orig_work_uri:
                            raise ValueError(f"No cellar entry found for {celex}")
                        manif_uri = find_pdf_manifestation(client, orig_work_uri)
                        if not manif_uri:
                            raise ValueError(
                                f"No PDF manifestation found for original {celex}"
                            )
                        actual_celex = orig_celex or celex
                        if not cons_celex:
                            console.log(f"  [yellow]no consolidated[/yellow] for {celex}, using original")

                    # --- Step 2: idempotency check ---
                    local_filename = f"celex_{actual_celex}_{short_name}.pdf"
                    local_path = DOCS_PRIMARY / local_filename

                    existing = index.get(actual_celex) or index.get(celex)
                    if local_path.exists() and existing and existing.get("sha256"):
                        current_hash = sha256_of(local_path)
                        if current_hash == existing["sha256"]:
                            console.log(f"  [dim]unchanged[/dim] {local_filename}")
                            index[actual_celex] = dict(existing, status="unchanged")
                            unchanged += 1
                            progress.advance(task)
                            continue

                    # --- Step 3: download PDF (tries DOC_1, DOC_2, content-neg) ---
                    content = download_pdf_from_manifestation(client, manif_uri)

                    local_path.write_bytes(content)
                    file_hash = sha256_of(local_path)

                    row = {
                        "id": actual_celex,
                        "source": "eurlex",
                        "short_name": short_name,
                        "title": "",
                        "doc_type": doc_type_from_celex(celex),
                        "url": manif_uri,
                        "published_date": "",
                        "local_path": str(local_path.relative_to(BASE_DIR)),
                        "sha256": file_hash,
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "ok",
                    }
                    index[actual_celex] = row
                    console.log(f"  [green]ok[/green] {local_filename} ({len(content):,} bytes)")
                    new_count += 1

                except Exception as exc:
                    console.log(f"  [red]FAILED[/red] {celex}: {exc}")
                    index[celex] = {
                        "id": celex,
                        "source": "eurlex",
                        "short_name": short_name,
                        "title": "",
                        "doc_type": doc_type_from_celex(celex),
                        "url": "",
                        "published_date": "",
                        "local_path": "",
                        "sha256": "",
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "failed",
                    }
                    failed += 1

                progress.advance(task)

    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# EBA helpers
# ---------------------------------------------------------------------------
_EBA_REF_RE = re.compile(
    r'EBA[-/_ ]([A-Z]{2,4})[-/_ ](\d{4})[-/_ ](\d{1,3})',
    re.IGNORECASE,
)


def _eba_ref_from_text(text: str) -> str | None:
    """Extract normalised EBA reference ID from any text, e.g. EBA/GL/2021/03."""
    m = _EBA_REF_RE.search(text)
    if m:
        kind = m.group(1).upper()
        year = m.group(2)
        num = m.group(3).zfill(2)
        return f"EBA/{kind}/{year}/{num}"
    return None


def _title_to_slug(title: str, max_len: int = 60) -> str:
    """Convert a title string to a lower-case underscore slug, max max_len chars."""
    slug = re.sub(r'[^a-z0-9]+', '_', title.lower().strip())
    slug = slug.strip('_')
    return slug[:max_len].rstrip('_')


def _eba_slug_from_ref(ref: str) -> str:
    """'EBA/GL/2021/03' → 'eba_gl_2021_03'"""
    return re.sub(r'[^a-z0-9]+', '_', ref.lower()).strip('_')


def _parse_eba_date(date_str: str) -> str:
    """Parse '2 March 2026' → '2026-03-02', or return '' on failure."""
    for fmt in ("%d %B %Y", "%B %d, %Y", "%d %b %Y"):
        try:
            return datetime.strptime(date_str.strip(), fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _get_last_page(soup: BeautifulSoup) -> int:
    """Return the 0-based index of the last page from pager, or 0 if only one page."""
    last = soup.find("li", class_=lambda x: x and "pager__item--last" in x)
    if last:
        a = last.find("a")
        if a and a.get("href"):
            m = re.search(r'page=(\d+)', a["href"])
            if m:
                return int(m.group(1))
    return 0


def _scrape_eba_listing_page(soup: BeautifulSoup) -> list[dict]:
    """Extract items from one EBA listing page. Returns list of dicts."""
    items = []
    for art in soup.find_all("article", class_="teaser"):
        # Title and primary link
        h4 = art.find("h4")
        if not h4:
            continue
        title_a = h4.find("a")
        if not title_a:
            continue
        title = title_a.get_text(strip=True)
        pdf_href = title_a.get("href", "")

        # Date
        cal = art.find(class_="link-icon--calendar")
        date_raw = cal.get_text(strip=True) if cal else ""

        items.append({
            "title": title,
            "pdf_href": pdf_href,
            "date_raw": date_raw,
        })
    return items


# ---------------------------------------------------------------------------
# EBA downloader
# ---------------------------------------------------------------------------
def download_eba(index: dict[str, dict]) -> tuple[int, int, int]:
    """
    Crawl EBA guidelines, recommendations, and opinions listing pages.
    Download the primary PDF for each item.
    Returns (new_count, unchanged, failed).
    """
    DOCS_EBA.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    _eba_logger.info("=== EBA crawl started ===")

    with httpx.Client(headers=headers, follow_redirects=True, timeout=120) as client:
        for doc_type, type_id in EBA_CATEGORIES.items():
            _eba_logger.info("Category: %s (type_id=%s)", doc_type, type_id)
            console.rule(f"[bold cyan]EBA {doc_type.capitalize()}s[/bold cyan]")

            listing_url = (
                f"{EBA_BASE}/publications-and-media/publications"
                f"?text=&document_type={type_id}&media_topics=All"
            )

            # --- Discover total pages ---
            try:
                resp = _get(client, listing_url)
                resp.raise_for_status()
            except Exception as exc:
                _eba_logger.error("Failed to fetch listing page 0 for %s: %s", doc_type, exc)
                console.log(f"[red]FAILED[/red] listing page 0 for {doc_type}: {exc}")
                continue

            soup = BeautifulSoup(resp.text, "lxml")
            last_page = _get_last_page(soup)
            total_pages = last_page + 1
            _eba_logger.info("  %s: %d pages", doc_type, total_pages)
            console.log(f"  {doc_type}: {total_pages} page(s) to crawl")

            # Collect all items across all pages
            all_items: list[dict] = []
            all_items.extend(_scrape_eba_listing_page(soup))

            for page_idx in range(1, total_pages):
                page_url = f"{listing_url}&page={page_idx}"
                try:
                    resp = _get(client, page_url)
                    resp.raise_for_status()
                    page_soup = BeautifulSoup(resp.text, "lxml")
                    page_items = _scrape_eba_listing_page(page_soup)
                    all_items.extend(page_items)
                    _eba_logger.info("  page %d: %d items", page_idx + 1, len(page_items))
                except Exception as exc:
                    _eba_logger.error("Failed page %d for %s: %s", page_idx + 1, doc_type, exc)
                    console.log(f"[red]FAILED[/red] page {page_idx + 1}: {exc}")

            _eba_logger.info("  %s total items found: %d", doc_type, len(all_items))
            console.log(f"  {doc_type}: {len(all_items)} items found")

            # --- Download each item ---
            slug_seen: set[str] = set()

            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(),
                TaskProgressColumn(),
                console=console,
            ) as progress:
                task = progress.add_task(
                    f"Downloading EBA {doc_type}s...", total=len(all_items)
                )

                for item in all_items:
                    title = item["title"]
                    pdf_href = item["pdf_href"]
                    date_raw = item["date_raw"]

                    progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")

                    if not pdf_href:
                        _eba_logger.warning("No PDF href for: %s", title)
                        progress.advance(task)
                        continue

                    # Build absolute URL
                    if pdf_href.startswith("/"):
                        pdf_url = EBA_BASE + pdf_href
                    elif pdf_href.startswith("http"):
                        pdf_url = pdf_href
                    else:
                        pdf_url = EBA_BASE + "/" + pdf_href

                    # Determine EBA ref ID from title or filename
                    eba_ref = _eba_ref_from_text(title) or _eba_ref_from_text(pdf_href)

                    if eba_ref:
                        slug = _eba_slug_from_ref(eba_ref)
                        doc_id = eba_ref
                    else:
                        slug = "eba_" + _title_to_slug(title)
                        doc_id = f"eba:{slug}"

                    # Ensure no collisions
                    base_slug = slug
                    counter = 1
                    while slug in slug_seen:
                        slug = f"{base_slug}_{counter}"
                        counter += 1
                    slug_seen.add(slug)

                    local_path = DOCS_EBA / f"{slug}.pdf"
                    published_date = _parse_eba_date(date_raw)

                    # --- Idempotency check ---
                    existing = index.get(doc_id)
                    if local_path.exists() and existing and existing.get("sha256"):
                        current_hash = sha256_of(local_path)
                        if current_hash == existing["sha256"]:
                            index[doc_id] = dict(existing, status="unchanged")
                            upsert_index_row(index[doc_id])
                            unchanged += 1
                            _eba_logger.debug("unchanged: %s", slug)
                            progress.advance(task)
                            continue

                    # --- Download PDF ---
                    try:
                        resp = _get(client, pdf_url)
                        resp.raise_for_status()
                        content = resp.content

                        if not content.startswith(b"%PDF"):
                            raise ValueError(
                                f"Response is not a PDF (first bytes: {content[:20]!r})"
                            )

                        local_path.write_bytes(content)
                        file_hash = sha256_of(local_path)

                        row = {
                            "id": doc_id,
                            "source": "eba",
                            "short_name": slug,
                            "title": title,
                            "doc_type": doc_type,
                            "url": pdf_url,
                            "published_date": published_date,
                            "local_path": str(local_path.relative_to(BASE_DIR)),
                            "sha256": file_hash,
                            "downloaded_at": datetime.now(timezone.utc).isoformat(),
                            "status": "ok",
                        }
                        index[doc_id] = row
                        upsert_index_row(row)
                        console.log(f"  [green]ok[/green] {slug}.pdf ({len(content):,} bytes)")
                        _eba_logger.info("ok: %s  %s  (%d bytes)", slug, pdf_url, len(content))
                        new_count += 1

                    except Exception as exc:
                        _eba_logger.error("FAILED %s: %s", slug, exc)
                        console.log(f"  [red]FAILED[/red] {slug}: {exc}")
                        row = {
                            "id": doc_id,
                            "source": "eba",
                            "short_name": slug,
                            "title": title,
                            "doc_type": doc_type,
                            "url": pdf_url,
                            "published_date": published_date,
                            "local_path": "",
                            "sha256": "",
                            "downloaded_at": datetime.now(timezone.utc).isoformat(),
                            "status": "failed",
                        }
                        index[doc_id] = row
                        upsert_index_row(row)
                        failed += 1

                    progress.advance(task)

            _eba_logger.info(
                "Category %s done — new=%d unchanged=%d failed=%d",
                doc_type, new_count, unchanged, failed,
            )

    _eba_logger.info(
        "=== EBA crawl finished — new=%d unchanged=%d failed=%d ===",
        new_count, unchanged, failed,
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# Shared slug helpers (used by ECB; EBA keeps its own inline set)
# ---------------------------------------------------------------------------

def _load_existing_slugs(target_dir: Path, index: dict[str, dict], source: str) -> set[str]:
    """
    Build a set of slugs already committed on disk or in index.csv for *source*.
    This prevents re-run suffix inflation (_1, _2 …) for already-downloaded docs.
    """
    slugs: set[str] = set()
    # From index.csv rows
    for row in index.values():
        if row.get("source") == source and row.get("short_name"):
            slugs.add(row["short_name"])
    # From files already on disk (catches renames / manual copies)
    for p in target_dir.glob("*"):
        slugs.add(p.stem)
    return slugs


def _unique_slug(base_slug: str, slug_seen: set[str]) -> str:
    """Return base_slug if unused, else base_slug_1, base_slug_2 …"""
    slug = base_slug
    counter = 1
    while slug in slug_seen:
        slug = f"{base_slug}_{counter}"
        counter += 1
    slug_seen.add(slug)
    return slug


# ---------------------------------------------------------------------------
# ECB helpers
# ---------------------------------------------------------------------------

def _ecb_foedb_load_all(client: httpx.Client) -> list[dict]:
    """
    Download all records from the ECB Banking Supervision publications foedb.
    Returns list of dicts keyed by ECB_FOEDB_HEADER.
    """
    # 1. Discover current version + hash
    versions_url = f"{ECB_BASE}{ECB_FOEDB_HOST}/{ECB_FOEDB_DB}/versions.json"
    resp = _get(client, versions_url)
    resp.raise_for_status()
    versions = resp.json()
    if not versions:
        raise ValueError("foedb versions.json returned empty list")
    version = versions[0]["version"]
    hash_ = versions[0]["hash"]
    _ecb_logger.info("foedb version=%s hash=%s", version, hash_)

    # 2. Fetch metadata
    meta_url = f"{ECB_BASE}{ECB_FOEDB_HOST}/{ECB_FOEDB_DB}/{version}/{hash_}/metadata.json"
    meta = _get(client, meta_url).json()
    _ecb_logger.info("foedb total_records=%d chunk_size=%d", meta["total_records"], meta["chunk_size"])

    # 3. Download chunks
    N = len(ECB_FOEDB_HEADER)
    all_records: list[dict] = []
    chunk_idx = 0
    while True:
        chunk_url = f"{ECB_BASE}{ECB_FOEDB_HOST}/{ECB_FOEDB_DB}/{version}/{hash_}/data/0/chunk_{chunk_idx}.json"
        resp = _get(client, chunk_url)
        if resp.status_code == 404:
            break
        resp.raise_for_status()
        flat = resp.json()
        records = [dict(zip(ECB_FOEDB_HEADER, flat[i:i + N])) for i in range(0, len(flat), N)]
        all_records.extend(records)
        _ecb_logger.debug("foedb chunk_%d: %d records", chunk_idx, len(records))
        chunk_idx += 1

    _ecb_logger.info("foedb loaded %d total records across %d chunks", len(all_records), chunk_idx)
    return all_records


def _ecb_parse_date(ts: int | None) -> str:
    """Unix timestamp → ISO date string, or ''."""
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(ts, tz=timezone.utc).date().isoformat()
    except Exception:
        return ""


def _ecb_pick_en_doc(doc_types: list | None) -> str | None:
    """
    From a record's documentTypes list, pick the best English document URL:
    prefer .en.pdf, fall back to .en.html, ignore other languages/filetypes.
    Returns None if nothing usable found.
    """
    if not doc_types:
        return None
    en_pdfs = [d for d in doc_types if isinstance(d, str) and d.endswith(".en.pdf")]
    if en_pdfs:
        return en_pdfs[0]
    # PDF with query string (cache-busting hash)
    en_pdfs_qs = [d for d in doc_types if isinstance(d, str) and ".en.pdf" in d and d.startswith("/")]
    if en_pdfs_qs:
        return en_pdfs_qs[0].split("?")[0]
    en_htmls = [d for d in doc_types if isinstance(d, str) and ".en.html" in d]
    if en_htmls:
        return en_htmls[0]
    return None


def _ecb_slug_from_url(url_path: str) -> str:
    """
    Derive a human-readable slug from an ECB URL path.
    e.g. '/ecb/pub/pdf/ssm.supervisory_guides202507.en.pdf' → 'ecb_supervisory_guides202507'
    """
    # Strip path components, extension, language suffix (.en)
    name = url_path.rstrip("/").split("/")[-1]
    name = re.sub(r"\?.*", "", name)          # remove query string
    name = re.sub(r"\.[a-z]{2}\.(pdf|html|xlsx)$", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\.(pdf|html|xlsx)$", "", name, flags=re.IGNORECASE)
    # Remove leading 'ssm.' prefix if present
    name = re.sub(r"^ssm\.", "", name, flags=re.IGNORECASE)
    # Normalise to slug chars
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    slug = slug[:60].rstrip("_")
    return f"ecb_{slug}"


def _ecb_download_doc(
    client: httpx.Client,
    url_path: str,
    slug: str,
) -> tuple[Path, str]:
    """
    Download document from ECB. url_path is relative (/ecb/pub/…) or absolute.
    Returns (local_path, sha256_hex). Raises on failure.
    """
    if url_path.startswith("http"):
        full_url = url_path
    else:
        full_url = ECB_BASE + url_path

    # Remove query-string cache-busters to get a clean URL
    full_url = full_url.split("?")[0]

    resp = _get(client, full_url)
    resp.raise_for_status()

    content_type = resp.headers.get("content-type", "")
    is_pdf = resp.content[:4] == b"%PDF"
    is_html = not is_pdf and ("html" in content_type or url_path.endswith(".html"))

    if is_pdf:
        local_path = DOCS_ECB / f"{slug}.pdf"
    else:
        local_path = DOCS_ECB / f"{slug}.html"

    local_path.write_bytes(resp.content)
    file_hash = sha256_of(local_path)
    return local_path, file_hash


def _ecb_process_record(
    client: httpx.Client,
    rec: dict,
    doc_type_label: str,
    index: dict[str, dict],
    slug_seen: set[str],
) -> tuple[str, str]:
    """
    Process one foedb record: download if needed, update index.
    Returns ('ok'|'unchanged'|'failed'|'skip', slug).
    """
    title = rec["publicationProperties"].get("Title", "") if isinstance(rec["publicationProperties"], dict) else ""
    doc_list = rec["documentTypes"]
    published_date = _ecb_parse_date(rec["pub_timestamp"])

    url_path = _ecb_pick_en_doc(doc_list)
    if not url_path:
        _ecb_logger.warning("No English document found for record id=%s title=%s", rec["id"], title[:60])
        return "skip", ""

    # Build slug from URL
    base_slug = _ecb_slug_from_url(url_path)
    # Check if we already have this slug in index (idempotency across runs)
    slug = _unique_slug(base_slug, slug_seen)
    doc_id = f"ecb:{slug}"

    # Determine final file extension
    is_pdf_url = ".en.pdf" in url_path or url_path.endswith(".pdf")
    ext = ".pdf" if is_pdf_url else ".html"
    local_path = DOCS_ECB / f"{slug}{ext}"

    # --- Idempotency check ---
    existing = index.get(doc_id)
    if local_path.exists() and existing and existing.get("sha256"):
        current_hash = sha256_of(local_path)
        if current_hash == existing["sha256"]:
            index[doc_id] = dict(existing, status="unchanged")
            upsert_index_row(index[doc_id])
            _ecb_logger.debug("unchanged: %s", slug)
            return "unchanged", slug

    # --- Download ---
    try:
        full_url = (ECB_BASE + url_path.split("?")[0]) if not url_path.startswith("http") else url_path.split("?")[0]
        actual_path, file_hash = _ecb_download_doc(client, url_path, slug)

        row = {
            "id": doc_id,
            "source": "ecb",
            "short_name": slug,
            "title": title,
            "doc_type": doc_type_label,
            "url": full_url,
            "published_date": published_date,
            "local_path": str(actual_path.relative_to(BASE_DIR)),
            "sha256": file_hash,
            "downloaded_at": datetime.now(timezone.utc).isoformat(),
            "status": "ok",
        }
        index[doc_id] = row
        upsert_index_row(row)
        _ecb_logger.info("ok: %s  %s  (%s)", slug, full_url, published_date)
        return "ok", slug

    except Exception as exc:
        _ecb_logger.error("FAILED %s  %s: %s", slug, url_path, exc)
        row = {
            "id": doc_id,
            "source": "ecb",
            "short_name": slug,
            "title": title,
            "doc_type": doc_type_label,
            "url": url_path,
            "published_date": published_date,
            "local_path": "",
            "sha256": "",
            "downloaded_at": datetime.now(timezone.utc).isoformat(),
            "status": "failed",
        }
        index[doc_id] = row
        upsert_index_row(row)
        return "failed", slug


def _ecb_scrape_letters_to_banks(
    client: httpx.Client,
    index: dict[str, dict],
    slug_seen: set[str],
) -> tuple[int, int, int]:
    """
    Scrape the ECB letters-to-banks listing page (static HTML, DL/DT/DD structure).
    Returns (new_count, unchanged, failed).
    """
    new_count = unchanged = failed = 0
    url = f"{ECB_BASE}/activities/letterstobanks/html/index.en.html"
    _ecb_logger.info("Scraping letters-to-banks: %s", url)

    try:
        resp = _get(client, url)
        resp.raise_for_status()
    except Exception as exc:
        _ecb_logger.error("Failed to fetch letters-to-banks listing: %s", exc)
        return 0, 0, 1

    soup = BeautifulSoup(resp.text, "lxml")

    items: list[dict] = []
    for dl in soup.find_all("dl"):
        current_date = ""
        for child in dl.children:
            if not hasattr(child, "name"):
                continue
            if child.name == "dt":
                text = child.get_text(strip=True)
                if text:
                    current_date = text
            elif child.name == "dd":
                title_div = child.find("div", class_="title")
                if not title_div:
                    continue
                # Try to get the English PDF href
                en_pdf_a = title_div.find("a", href=lambda h: h and ".en.pdf" in h)
                # Fall back to English HTML
                en_html_a = title_div.find("a", href=lambda h: h and ".en.html" in h)
                primary_a = title_div.find("a")
                if not primary_a:
                    continue
                title = primary_a.get_text(strip=True)
                href = (en_pdf_a or en_html_a or primary_a).get("href", "")
                if not href:
                    continue
                items.append({"title": title, "href": href.split("?")[0], "date_raw": current_date})

    _ecb_logger.info("letters-to-banks: found %d items", len(items))
    console.log(f"  letters-to-banks: {len(items)} items found")

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Downloading ECB letters...", total=len(items))

        for item in items:
            title = item["title"]
            href = item["href"]
            progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")

            # Parse date
            pub_date = ""
            for fmt in ("%d %B %Y", "%B %d, %Y", "%d/%m/%Y", "%d %b %Y"):
                try:
                    pub_date = datetime.strptime(item["date_raw"].strip(), fmt).date().isoformat()
                    break
                except ValueError:
                    pass

            base_slug = _ecb_slug_from_url(href)
            slug = _unique_slug(base_slug, slug_seen)
            doc_id = f"ecb:{slug}"

            is_pdf_url = href.endswith(".pdf")
            ext = ".pdf" if is_pdf_url else ".html"
            local_path = DOCS_ECB / f"{slug}{ext}"

            existing = index.get(doc_id)
            if local_path.exists() and existing and existing.get("sha256"):
                current_hash = sha256_of(local_path)
                if current_hash == existing["sha256"]:
                    index[doc_id] = dict(existing, status="unchanged")
                    upsert_index_row(index[doc_id])
                    unchanged += 1
                    _ecb_logger.debug("unchanged: %s", slug)
                    progress.advance(task)
                    continue

            try:
                actual_path, file_hash = _ecb_download_doc(client, href, slug)
                full_url = (ECB_BASE + href) if href.startswith("/") else href
                row = {
                    "id": doc_id,
                    "source": "ecb",
                    "short_name": slug,
                    "title": title,
                    "doc_type": "letter",
                    "url": full_url,
                    "published_date": pub_date,
                    "local_path": str(actual_path.relative_to(BASE_DIR)),
                    "sha256": file_hash,
                    "downloaded_at": datetime.now(timezone.utc).isoformat(),
                    "status": "ok",
                }
                index[doc_id] = row
                upsert_index_row(row)
                console.log(f"  [green]ok[/green] {slug}{ext}")
                _ecb_logger.info("ok: %s  %s", slug, full_url)
                new_count += 1
            except Exception as exc:
                _ecb_logger.error("FAILED %s: %s", slug, exc)
                console.log(f"  [red]FAILED[/red] {slug}: {exc}")
                row = {
                    "id": doc_id, "source": "ecb", "short_name": slug,
                    "title": title, "doc_type": "letter", "url": href,
                    "published_date": pub_date, "local_path": "", "sha256": "",
                    "downloaded_at": datetime.now(timezone.utc).isoformat(),
                    "status": "failed",
                }
                index[doc_id] = row
                upsert_index_row(row)
                failed += 1

            progress.advance(task)

    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# ECB downloader
# ---------------------------------------------------------------------------
def download_ecb(index: dict[str, dict]) -> tuple[int, int, int]:
    """
    Download ECB Banking Supervision documents:
    1. Supervisory Guides (foedb type 305)
    2. Dear-CEO letters / letters to banks (static HTML scrape)
    3. Supervisory Priorities (foedb type 186)
    4. Other publications (foedb, all remaining doc types from ECB_MIN_YEAR+)
    Returns (new_count, unchanged, failed).
    """
    DOCS_ECB.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    _ecb_logger.info("=== ECB crawl started ===")

    # Pre-load slug set from existing files + index to avoid suffix inflation on re-runs
    slug_seen: set[str] = _load_existing_slugs(DOCS_ECB, index, "ecb")

    with httpx.Client(headers=headers, follow_redirects=True, timeout=120) as client:
        # ------------------------------------------------------------------
        # Step 1: Load all foedb records once
        # ------------------------------------------------------------------
        console.rule("[bold cyan]ECB Banking Supervision[/bold cyan]")
        try:
            all_records = _ecb_foedb_load_all(client)
        except Exception as exc:
            _ecb_logger.error("Failed to load foedb: %s", exc)
            console.log(f"[red]FAILED[/red] to load ECB foedb: {exc}")
            # Still attempt static-HTML crawls
            all_records = []

        # ------------------------------------------------------------------
        # Step 2: Supervisory Guides
        # ------------------------------------------------------------------
        console.rule("[cyan]ECB Supervisory Guides[/cyan]")
        guide_recs = [r for r in all_records if r["type"] == ECB_TYPE_GUIDE]
        _ecb_logger.info("Guides: %d records", len(guide_recs))
        console.log(f"  Guides: {len(guide_recs)} records")

        with Progress(
            SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(), console=console,
        ) as progress:
            task = progress.add_task("Downloading ECB Guides...", total=len(guide_recs))
            for rec in guide_recs:
                title = (rec["publicationProperties"] or {}).get("Title", "?") if isinstance(rec.get("publicationProperties"), dict) else "?"
                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")
                status, slug = _ecb_process_record(client, rec, "guide", index, slug_seen)
                if status == "ok":
                    new_count += 1
                    console.log(f"  [green]ok[/green] {slug}")
                elif status == "unchanged":
                    unchanged += 1
                elif status == "failed":
                    failed += 1
                    console.log(f"  [red]FAILED[/red] {title[:50]}")
                progress.advance(task)

        _ecb_logger.info("Guides done — new=%d unchanged=%d failed=%d", new_count, unchanged, failed)

        # ------------------------------------------------------------------
        # Step 3: Letters to banks (static HTML)
        # ------------------------------------------------------------------
        console.rule("[cyan]ECB Letters to Banks[/cyan]")
        n, u, f = _ecb_scrape_letters_to_banks(client, index, slug_seen)
        new_count += n
        unchanged += u
        failed += f
        _ecb_logger.info("Letters done — new=%d unchanged=%d failed=%d", n, u, f)

        # ------------------------------------------------------------------
        # Step 4: Supervisory Priorities
        # ------------------------------------------------------------------
        console.rule("[cyan]ECB Supervisory Priorities[/cyan]")
        prio_recs = [
            r for r in all_records
            if r["type"] == ECB_TYPE_PRIORITIES and r.get("year", 0) >= ECB_MIN_YEAR
        ]
        _ecb_logger.info("Priorities: %d records", len(prio_recs))
        console.log(f"  Priorities: {len(prio_recs)} records")

        with Progress(
            SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(), console=console,
        ) as progress:
            task = progress.add_task("Downloading ECB Priorities...", total=len(prio_recs))
            pn = pu = pf = 0
            for rec in prio_recs:
                title = (rec["publicationProperties"] or {}).get("Title", "?") if isinstance(rec.get("publicationProperties"), dict) else "?"
                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")
                status, slug = _ecb_process_record(client, rec, "priorities", index, slug_seen)
                if status == "ok":
                    new_count += 1; pn += 1
                    console.log(f"  [green]ok[/green] {slug}")
                elif status == "unchanged":
                    unchanged += 1; pu += 1
                elif status == "failed":
                    failed += 1; pf += 1
                    console.log(f"  [red]FAILED[/red] {title[:50]}")
                progress.advance(task)

        _ecb_logger.info("Priorities done — new=%d unchanged=%d failed=%d", pn, pu, pf)

        # ------------------------------------------------------------------
        # Step 5: General publications (catch-all, exclude speeches/interviews etc.)
        # ------------------------------------------------------------------
        console.rule("[cyan]ECB Publications (general)[/cyan]")
        pub_recs = [
            r for r in all_records
            if r["type"] not in ECB_PUB_EXCLUDE_TYPES
            and r["type"] not in {ECB_TYPE_GUIDE, ECB_TYPE_PRIORITIES}
            and r.get("year", 0) >= ECB_MIN_YEAR
        ]
        _ecb_logger.info("Publications (general): %d records", len(pub_recs))
        console.log(f"  Publications (general): {len(pub_recs)} records")

        with Progress(
            SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(), console=console,
        ) as progress:
            task = progress.add_task("Downloading ECB Publications...", total=len(pub_recs))
            qn = qu = qf = 0
            for rec in pub_recs:
                title = (rec["publicationProperties"] or {}).get("Title", "?") if isinstance(rec.get("publicationProperties"), dict) else "?"
                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")
                status, slug = _ecb_process_record(client, rec, "publication", index, slug_seen)
                if status == "ok":
                    new_count += 1; qn += 1
                    console.log(f"  [green]ok[/green] {slug}")
                elif status == "unchanged":
                    unchanged += 1; qu += 1
                elif status == "failed":
                    failed += 1; qf += 1
                    console.log(f"  [red]FAILED[/red] {title[:50]}")
                progress.advance(task)

        _ecb_logger.info("Publications done — new=%d unchanged=%d failed=%d", qn, qu, qf)

    _ecb_logger.info(
        "=== ECB crawl finished — new=%d unchanged=%d failed=%d ===",
        new_count, unchanged, failed,
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# ESMA helpers
# ---------------------------------------------------------------------------

def _esma_ref_from_url(url: str) -> str | None:
    """
    Extract a normalised ESMA document reference from a URL or filename.
    Returns the matched group (original casing preserved) or None.
    e.g. '/sites/default/files/2025-02/ESMA35-1872330276-2032_Guidelines.pdf'
         → 'ESMA35-1872330276-2032'
    """
    basename = url.rstrip("/").split("/")[-1]
    # Remove .pdf extension
    basename = re.sub(r"\.pdf$", "", basename, flags=re.IGNORECASE)
    m = _ESMA_REF_RE.search(basename)
    return m.group(1) if m else None


def _esma_slug(ref_or_title: str, prefix: str = "esma") -> str:
    """
    Convert an ESMA ref (e.g. 'ESMA35-1872330276-2032') or a title string
    into a lower-case underscore slug, max 64 chars.
    Always prefixed with 'esma_' (unless ref already starts with it).
    """
    slug = re.sub(r"[^a-z0-9]+", "_", ref_or_title.lower()).strip("_")
    slug = slug[:64].rstrip("_")
    if not slug.startswith("esma_"):
        slug = "esma_" + slug
    return slug


def _esma_is_lang_suffix(basename_no_ext: str) -> bool:
    """Return True if basename ends with a 2-letter EU language code (non-English)."""
    m = re.search(r"_([A-Z]{2})$", basename_no_ext, re.IGNORECASE)
    if m:
        return m.group(1).upper() in _ESMA_LANG_SUFFIXES
    return False


def _esma_pick_en_pdf_from_doc_page(
    client: httpx.Client, doc_url: str
) -> str | None:
    """
    Fetch an ESMA /document/ page and return the English PDF href (relative),
    or None if not found.  English = no EU language suffix before .pdf.
    """
    try:
        resp = _get(client, doc_url)
        resp.raise_for_status()
    except Exception as exc:
        _esma_logger.warning("Failed to fetch doc page %s: %s", doc_url, exc)
        return None

    soup = BeautifulSoup(resp.text, "lxml")
    for a in soup.find_all("a", href=True):
        href: str = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        basename_no_ext = href.rstrip("/").split("/")[-1]
        basename_no_ext = re.sub(r"\.pdf$", "", basename_no_ext, flags=re.IGNORECASE)
        if not _esma_is_lang_suffix(basename_no_ext):
            return href
    return None


def _esma_parse_date(value) -> str:
    """
    Parse various date representations to ISO date string.
    Accepts datetime objects (from openpyxl) or string 'DD/MM/YYYY'.
    """
    if value is None:
        return ""
    if hasattr(value, "date"):
        try:
            return value.date().isoformat()
        except Exception:
            return ""
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _esma_load_guidelines_from_tracker(
    client: httpx.Client,
) -> list[dict]:
    """
    Download and parse the ESMA guidelines_tracker.xlsx.
    Returns a list of dicts with keys:
      title, directive, url, date_raw, url_type ('direct_pdf'|'doc_page'|'skip')
    """
    try:
        resp = _get(client, ESMA_GUIDELINES_TRACKER_URL)
        resp.raise_for_status()
    except Exception as exc:
        _esma_logger.error("Failed to download guidelines tracker: %s", exc)
        return []

    # openpyxl is a soft dependency; import here so the module still loads without it
    try:
        import openpyxl
    except ImportError:
        _esma_logger.error("openpyxl not installed; cannot parse guidelines tracker")
        return []

    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
    ws = wb["Guidelines Tracker"]

    items: list[dict] = []
    # Row 1 is a merged header, row 2 is column names, data starts row 3
    for row_idx in range(3, ws.max_row + 1):
        title_cell = ws.cell(row=row_idx, column=2)
        directive_cell = ws.cell(row=row_idx, column=1)
        # Column 7 = Translation in official languages (English text)
        translation_cell = ws.cell(row=row_idx, column=7)
        # Column 8 = Date of application
        date_cell = ws.cell(row=row_idx, column=8)

        title = title_cell.value
        if not title:
            continue
        title = str(title).strip()
        if not title or title.lower().startswith("esma guidelines"):
            # Skip header rows
            continue

        directive = str(directive_cell.value or "").strip()
        date_raw = date_cell.value  # datetime or string

        url: str | None = None
        url_type: str = "skip"

        if translation_cell.hyperlink:
            raw_url = translation_cell.hyperlink.target or ""
            if raw_url.endswith(".pdf") or ".pdf" in raw_url:
                # May contain a double URL in old entries; take the first
                clean = raw_url.split("https:/", 1)
                if len(clean) > 1:
                    url = "https:/" + clean[1].split("https:/")[0]
                else:
                    url = raw_url
                url_type = "direct_pdf"
            elif "/document/" in raw_url:
                url = raw_url
                url_type = "doc_page"
            else:
                # databases-library, press-news, external EBA etc. → skip
                url = raw_url
                url_type = "skip"
        else:
            url_type = "skip"

        items.append({
            "title": title,
            "directive": directive,
            "url": url or "",
            "url_type": url_type,
            "date_raw": date_raw,
        })

    _esma_logger.info(
        "Guidelines tracker: %d rows parsed — direct_pdf=%d doc_page=%d skip=%d",
        len(items),
        sum(1 for i in items if i["url_type"] == "direct_pdf"),
        sum(1 for i in items if i["url_type"] == "doc_page"),
        sum(1 for i in items if i["url_type"] == "skip"),
    )
    return items


def _esma_load_qa_bundles(client: httpx.Client) -> list[dict]:
    """
    Scrape the ESMA Q&As listing page and return a list of dicts:
      { title, url }  where url is an absolute PDF URL.
    """
    try:
        resp = _get(client, ESMA_QA_BUNDLES_URL)
        resp.raise_for_status()
    except Exception as exc:
        _esma_logger.error("Failed to fetch Q&A bundles page: %s", exc)
        return []

    soup = BeautifulSoup(resp.text, "lxml")
    items: list[dict] = []
    seen_hrefs: set[str] = set()

    for a in soup.find_all("a", href=True):
        href: str = a["href"]
        if not href.lower().endswith(".pdf"):
            continue
        if href in seen_hrefs:
            continue
        seen_hrefs.add(href)

        title = a.get_text(strip=True)
        if not title:
            continue

        full_url = (ESMA_BASE + href) if href.startswith("/") else href
        items.append({"title": title, "url": full_url})

    _esma_logger.info("Q&A bundles page: %d PDF links found", len(items))
    return items


# ---------------------------------------------------------------------------
# ESMA downloader
# ---------------------------------------------------------------------------

def download_esma(index: dict[str, dict]) -> tuple[int, int, int]:
    """
    Crawl ESMA Guidelines (via guidelines_tracker.xlsx) and Q&A bundle PDFs.
    Returns (new_count, unchanged, failed).
    """
    DOCS_ESMA.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }

    _esma_logger.info("=== ESMA crawl started ===")

    # Pre-load slug set to prevent suffix inflation on re-runs
    slug_seen: set[str] = _load_existing_slugs(DOCS_ESMA, index, "esma")

    with httpx.Client(headers=headers, follow_redirects=True, timeout=120) as client:

        # ---------------------------------------------------------------
        # Part 1: Guidelines from tracker Excel
        # ---------------------------------------------------------------
        console.rule("[bold cyan]ESMA Guidelines[/bold cyan]")
        _esma_logger.info("Loading ESMA guidelines tracker …")

        guideline_items = _esma_load_guidelines_from_tracker(client)
        console.log(f"  Guidelines tracker: {len(guideline_items)} rows")

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task(
                "Downloading ESMA Guidelines…", total=len(guideline_items)
            )

            for item in guideline_items:
                title = item["title"]
                url_type = item["url_type"]
                raw_url = item["url"]
                date_raw = item["date_raw"]

                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")

                if url_type == "skip" or not raw_url:
                    _esma_logger.warning(
                        "SKIP (no downloadable link): %s — url=%s", title[:80], raw_url[:80]
                    )
                    progress.advance(task)
                    continue

                # Resolve actual PDF URL
                pdf_url: str | None = None
                if url_type == "direct_pdf":
                    pdf_url = (
                        (ESMA_BASE + raw_url) if raw_url.startswith("/") else raw_url
                    )
                elif url_type == "doc_page":
                    # Scrape the document page for the English PDF link
                    en_href = _esma_pick_en_pdf_from_doc_page(client, raw_url)
                    if not en_href:
                        _esma_logger.warning(
                            "SKIP (no English PDF on doc page): %s — %s", title[:80], raw_url
                        )
                        progress.advance(task)
                        continue
                    pdf_url = (ESMA_BASE + en_href) if en_href.startswith("/") else en_href

                if not pdf_url:
                    _esma_logger.warning("SKIP (pdf_url empty): %s", title[:80])
                    progress.advance(task)
                    continue

                # Derive ID / slug from PDF URL
                ref = _esma_ref_from_url(pdf_url)
                if ref:
                    base_slug = _esma_slug(ref)
                    doc_id = ref.upper()
                else:
                    base_slug = _esma_slug(title)
                    doc_id = f"esma:{base_slug}"

                slug = _unique_slug(base_slug, slug_seen)
                local_path = DOCS_ESMA / f"{slug}.pdf"
                published_date = _esma_parse_date(date_raw)

                # Idempotency check
                existing = index.get(doc_id)
                if local_path.exists() and existing and existing.get("sha256"):
                    current_hash = sha256_of(local_path)
                    if current_hash == existing["sha256"]:
                        index[doc_id] = dict(existing, status="unchanged")
                        upsert_index_row(index[doc_id])
                        unchanged += 1
                        _esma_logger.debug("unchanged: %s", slug)
                        progress.advance(task)
                        continue

                # Download PDF
                try:
                    resp = _get(client, pdf_url)
                    resp.raise_for_status()
                    content = resp.content

                    if not content.startswith(b"%PDF"):
                        raise ValueError(
                            f"Not a PDF (first bytes: {content[:20]!r})"
                        )

                    local_path.write_bytes(content)
                    file_hash = sha256_of(local_path)

                    row = {
                        "id": doc_id,
                        "source": "esma",
                        "short_name": slug,
                        "title": title,
                        "doc_type": "guideline",
                        "url": pdf_url,
                        "published_date": published_date,
                        "local_path": str(local_path.relative_to(BASE_DIR)),
                        "sha256": file_hash,
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "ok",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    console.log(f"  [green]ok[/green] {slug}.pdf ({len(content):,} bytes)")
                    _esma_logger.info("ok: %s  %s  (%d bytes)", slug, pdf_url, len(content))
                    new_count += 1

                except Exception as exc:
                    _esma_logger.error("FAILED %s: %s", slug, exc)
                    console.log(f"  [red]FAILED[/red] {slug}: {exc}")
                    row = {
                        "id": doc_id,
                        "source": "esma",
                        "short_name": slug,
                        "title": title,
                        "doc_type": "guideline",
                        "url": pdf_url,
                        "published_date": published_date,
                        "local_path": "",
                        "sha256": "",
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "failed",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    failed += 1

                progress.advance(task)

        _esma_logger.info(
            "Guidelines done — new=%d unchanged=%d failed=%d",
            new_count, unchanged, failed,
        )

        # ---------------------------------------------------------------
        # Part 2: Q&A bundle PDFs
        # ---------------------------------------------------------------
        console.rule("[bold cyan]ESMA Q&A Bundles[/bold cyan]")
        _esma_logger.info("Loading ESMA Q&A bundles …")

        qa_items = _esma_load_qa_bundles(client)
        console.log(f"  Q&A bundles page: {len(qa_items)} PDFs found")

        qn = qu = qf = 0

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task(
                "Downloading ESMA Q&As…", total=len(qa_items)
            )

            for item in qa_items:
                title = item["title"]
                pdf_url = item["url"]

                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")

                # Derive slug/ID
                ref = _esma_ref_from_url(pdf_url)
                if ref:
                    base_slug = _esma_slug(ref)
                    doc_id = ref.upper()
                else:
                    base_slug = _esma_slug(title)
                    doc_id = f"esma:{base_slug}"

                slug = _unique_slug(base_slug, slug_seen)
                local_path = DOCS_ESMA / f"{slug}.pdf"

                # Idempotency check
                existing = index.get(doc_id)
                if local_path.exists() and existing and existing.get("sha256"):
                    current_hash = sha256_of(local_path)
                    if current_hash == existing["sha256"]:
                        index[doc_id] = dict(existing, status="unchanged")
                        upsert_index_row(index[doc_id])
                        qu += 1
                        unchanged += 1
                        _esma_logger.debug("unchanged: %s", slug)
                        progress.advance(task)
                        continue

                try:
                    resp = _get(client, pdf_url)
                    resp.raise_for_status()
                    content = resp.content

                    if not content.startswith(b"%PDF"):
                        raise ValueError(f"Not a PDF (first bytes: {content[:20]!r})")

                    local_path.write_bytes(content)
                    file_hash = sha256_of(local_path)

                    row = {
                        "id": doc_id,
                        "source": "esma",
                        "short_name": slug,
                        "title": title,
                        "doc_type": "qa",
                        "url": pdf_url,
                        "published_date": "",
                        "local_path": str(local_path.relative_to(BASE_DIR)),
                        "sha256": file_hash,
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "ok",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    console.log(f"  [green]ok[/green] {slug}.pdf ({len(content):,} bytes)")
                    _esma_logger.info("ok: %s  %s  (%d bytes)", slug, pdf_url, len(content))
                    new_count += 1
                    qn += 1

                except Exception as exc:
                    _esma_logger.error("FAILED %s: %s", slug, exc)
                    console.log(f"  [red]FAILED[/red] {slug}: {exc}")
                    row = {
                        "id": doc_id,
                        "source": "esma",
                        "short_name": slug,
                        "title": title,
                        "doc_type": "qa",
                        "url": pdf_url,
                        "published_date": "",
                        "local_path": "",
                        "sha256": "",
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "failed",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    failed += 1
                    qf += 1

                progress.advance(task)

        _esma_logger.info(
            "Q&As done — new=%d unchanged=%d failed=%d", qn, qu, qf
        )

    _esma_logger.info(
        "=== ESMA crawl finished — new=%d unchanged=%d failed=%d ===",
        new_count, unchanged, failed,
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# FATF helpers
# ---------------------------------------------------------------------------

def _fatf_title_to_slug(title: str) -> str:
    """Convert a FATF doc title to a lowercase underscore slug, max 60 chars."""
    slug = re.sub(r"[^a-z0-9]+", "_", title.lower()).strip("_")
    slug = slug[:57].rstrip("_")
    return f"fatf_{slug}"


def _fatf_parse_date(ms: int | None) -> str:
    """Convert FATF API millisecond timestamp to ISO date string, or ''."""
    if not ms:
        return ""
    try:
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date().isoformat()
    except Exception:
        return ""


def _fatf_query_api(page, doc_type_filter: str, offset: int) -> dict:
    """
    Call the FATF AEM faceted-search API from inside the Playwright browser context.
    Returns the parsed JSON response dict (keys: results, totalMatches, limit, offset).
    """
    import urllib.parse
    encoded_filter = urllib.parse.quote(doc_type_filter)
    url = f"{FATF_SEARCH_API}?facet={encoded_filter}&offset={offset}"
    result = page.evaluate(
        f"""
        async () => {{
            const r = await fetch('{url}');
            return r.json();
        }}
        """
    )
    return result


def _fatf_listing_to_docs(page, doc_type_label: str, filter_value: str) -> list[dict]:
    """
    Paginate through the FATF search API for a given doc-type filter.
    Returns list of dicts: {{path, title, published_date, doc_type}}.
    Skips items whose category is in FATF_SKIP_CATEGORIES.
    """
    docs: list[dict] = []
    offset = 0
    total = None
    while True:
        data = _fatf_query_api(page, filter_value, offset)
        if total is None:
            total = data.get("totalMatches", 0)
            _fatf_logger.info(
                "facets.json filter=%r offset=%d totalMatches=%d",
                doc_type_label, offset, total,
            )
        results = data.get("results", [])
        if not results:
            break
        for item in results:
            path: str = item.get("path", "")
            # path = /content/fatf-gafi/en/publications/Category/doc-slug
            parts = path.strip("/").split("/")
            category = parts[-2] if len(parts) >= 2 else ""
            if category in FATF_SKIP_CATEGORIES:
                continue
            docs.append({
                "path": path,
                "title": item.get("title", ""),
                "published_date": _fatf_parse_date(item.get("publicationDate")),
                "doc_type": doc_type_label,
            })
        offset += len(results)
        if offset >= total:
            break
    _fatf_logger.info("filter=%r: %d docs (after skipping excluded categories)", doc_type_label, len(docs))
    return docs


def _fatf_pick_en_pdf(page, doc_html_url: str) -> str | None:
    """
    Load a FATF document HTML page and return the first English (non-translation) PDF URL.
    English PDFs live at /content/dam/fatf-gafi/ but NOT under /translations/.
    Returns full URL string or None.
    """
    try:
        resp = page.goto(doc_html_url, wait_until="networkidle", timeout=60000)
        if resp and resp.status == 404:
            _fatf_logger.warning("404 for doc page: %s", doc_html_url)
            return None
    except Exception as exc:
        _fatf_logger.warning("Failed to load doc page %s: %s", doc_html_url, exc)
        return None

    content = page.content()
    # Find all /content/dam/ PDF hrefs — the English one is NOT under /translations/
    pdf_paths = re.findall(
        r"""href=["'](/content/dam/fatf-gafi/[^"']*\.pdf[^"']*)["']""",
        content,
    )
    for href in pdf_paths:
        if "/translations/" not in href:
            return FATF_BASE + href
    # Fallback: take the very first PDF if all are translations
    if pdf_paths:
        return FATF_BASE + pdf_paths[0]
    return None


def _fatf_download_pdf(page, pdf_url: str, local_path: Path) -> str:
    """
    Download a PDF from a FATF URL using browser fetch (bypasses Cloudflare).
    Writes file to local_path and returns sha256 hex. Raises on failure.
    """
    result = page.evaluate(
        f"""
        async () => {{
            const r = await fetch('{pdf_url}');
            if (!r.ok) throw new Error('HTTP ' + r.status + ' for {pdf_url}');
            const buf = await r.arrayBuffer();
            const arr = new Uint8Array(buf);
            // Return as base64 string (JS btoa works on strings, use manual encoding)
            let binary = '';
            for (let i = 0; i < arr.length; i++) binary += String.fromCharCode(arr[i]);
            return btoa(binary);
        }}
        """
    )
    import base64
    data = base64.b64decode(result)
    if not data.startswith(b"%PDF"):
        raise ValueError(f"Response is not a PDF (first bytes: {data[:20]!r})")
    local_path.write_bytes(data)
    return sha256_of(local_path)


# ---------------------------------------------------------------------------
# FATF downloader
# ---------------------------------------------------------------------------

def download_fatf(index: dict[str, dict]) -> tuple[int, int, int]:
    """
    Crawl FATF standards & guidance corpus:
      - 40 Recommendations (Recommendations doc type)
      - Guidance papers (Guidance doc type)
      - Risk-Based Approach guidance (Risk Based Approach doc type)

    Uses Playwright throughout — the site is behind Cloudflare and blocks httpx.
    Transport: Playwright (all pages + PDF downloads via browser fetch).
    Returns (new_count, unchanged, failed).
    """
    from playwright.sync_api import sync_playwright  # soft import

    DOCS_FATF.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    _fatf_logger.info("=== FATF crawl started (transport: Playwright) ===")
    _fatf_logger.info(
        "NOTE: httpx blocked by Cloudflare on fatf-gafi.org; "
        "using Playwright browser for all requests."
    )

    slug_seen: set[str] = _load_existing_slugs(DOCS_FATF, index, "fatf")

    console.rule("[bold cyan]FATF Standards & Guidance[/bold cyan]")
    console.log("[yellow]Transport: Playwright (Cloudflare bypass required)[/yellow]")

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            user_agent=USER_AGENT,
            accept_downloads=True,
        )
        page = ctx.new_page()
        page.set_default_timeout(90000)

        # ------------------------------------------------------------------
        # Step 1: Navigate to the search page once so the browser gets
        #         Cloudflare clearance cookies, then query the API.
        # ------------------------------------------------------------------
        try:
            page.goto(
                f"{FATF_BASE}/en/search-page.html",
                wait_until="load",
                timeout=60000,
            )
            page.wait_for_timeout(2000)
            _fatf_logger.info("Search page loaded; Cloudflare cleared")
        except Exception as exc:
            _fatf_logger.error("Could not load FATF search page: %s", exc)
            console.log(f"[red]FAILED[/red] Could not load FATF search page: {exc}")
            browser.close()
            return 0, 0, 1

        # ------------------------------------------------------------------
        # Step 2: Discover all docs via the faceted-search API.
        # ------------------------------------------------------------------
        all_docs: list[dict] = []
        seen_paths: set[str] = set()

        for doc_type_label, filter_value in FATF_DOC_TYPE_FILTERS.items():
            try:
                docs = _fatf_listing_to_docs(page, doc_type_label, filter_value)
                # Dedup across doc types (same doc may match multiple filters)
                for d in docs:
                    if d["path"] not in seen_paths:
                        seen_paths.add(d["path"])
                        all_docs.append(d)
                _fatf_logger.info("filter=%s: %d unique docs added", doc_type_label, len(docs))
                console.log(f"  {doc_type_label}: {len(docs)} docs discovered")
            except Exception as exc:
                _fatf_logger.error("Failed to query API for filter=%s: %s", doc_type_label, exc)
                console.log(f"[red]FAILED[/red] API query for {doc_type_label}: {exc}")

        _fatf_logger.info("Total unique docs to process: %d", len(all_docs))
        console.log(f"  [bold]Total unique docs: {len(all_docs)}[/bold]")

        # ------------------------------------------------------------------
        # Step 3: For each doc, visit its HTML page, pick the English PDF,
        #         and download it.
        # ------------------------------------------------------------------
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("Downloading FATF docs...", total=len(all_docs))

            for doc in all_docs:
                path: str = doc["path"]
                title: str = doc["title"]
                doc_type: str = doc["doc_type"]
                published_date: str = doc["published_date"]

                # Convert content path → HTML URL
                html_path = path.replace("/content/fatf-gafi", "")
                doc_html_url = FATF_BASE + html_path + ".html"

                # Build slug + doc_id
                doc_slug_part = path.rstrip("/").split("/")[-1]
                base_slug = "fatf_" + re.sub(r"[^a-z0-9]+", "_", doc_slug_part.lower()).strip("_")
                base_slug = base_slug[:57].rstrip("_")

                # Check index for pre-existing slug to avoid suffix inflation
                existing_by_path = None
                for row in index.values():
                    if row.get("source") == "fatf" and row.get("url", "").find(doc_slug_part) != -1:
                        existing_by_path = row
                        base_slug = row["short_name"]
                        break

                slug = _unique_slug(base_slug, slug_seen)
                doc_id = f"fatf:{slug}"
                local_path = DOCS_FATF / f"{slug}.pdf"

                progress.update(task, description=f"[cyan]{title[:50]}[/cyan]")

                # --- Idempotency check ---
                existing = index.get(doc_id)
                if local_path.exists() and existing and existing.get("sha256"):
                    try:
                        current_hash = sha256_of(local_path)
                        if current_hash == existing["sha256"]:
                            index[doc_id] = dict(existing, status="unchanged")
                            upsert_index_row(index[doc_id])
                            unchanged += 1
                            _fatf_logger.debug("unchanged: %s", slug)
                            progress.advance(task)
                            continue
                    except Exception:
                        pass

                # --- Per-doc try/except: find PDF + download ---
                try:
                    # Rate-limit: throttle to ~1 req/s on this host
                    _throttle(FATF_BASE.split("//")[1])

                    pdf_url = _fatf_pick_en_pdf(page, doc_html_url)
                    if not pdf_url:
                        _fatf_logger.warning(
                            "No English PDF found for: %s  (%s)", title, doc_html_url
                        )
                        console.log(f"  [yellow]SKIP[/yellow] no English PDF: {title[:55]}")
                        progress.advance(task)
                        continue

                    file_hash = _fatf_download_pdf(page, pdf_url, local_path)

                    row = {
                        "id": doc_id,
                        "source": "fatf",
                        "short_name": slug,
                        "title": title,
                        "doc_type": doc_type,
                        "url": pdf_url,
                        "published_date": published_date,
                        "local_path": str(local_path.relative_to(BASE_DIR)),
                        "sha256": file_hash,
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "ok",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    sz = local_path.stat().st_size
                    console.log(f"  [green]ok[/green] {slug}.pdf ({sz:,} bytes)")
                    _fatf_logger.info("ok: %s  %s  (%d bytes)", slug, pdf_url, sz)
                    new_count += 1

                except Exception as exc:
                    _fatf_logger.error("FAILED %s  %s: %s", slug, doc_html_url, exc)
                    console.log(f"  [red]FAILED[/red] {slug}: {exc}")
                    row = {
                        "id": doc_id,
                        "source": "fatf",
                        "short_name": slug,
                        "title": title,
                        "doc_type": doc_type,
                        "url": doc_html_url,
                        "published_date": published_date,
                        "local_path": "",
                        "sha256": "",
                        "downloaded_at": datetime.now(timezone.utc).isoformat(),
                        "status": "failed",
                    }
                    index[doc_id] = row
                    upsert_index_row(row)
                    failed += 1

                progress.advance(task)

        browser.close()

    _fatf_logger.info(
        "=== FATF crawl finished — new=%d unchanged=%d failed=%d ===",
        new_count, unchanged, failed,
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# Irish Statute Book crawler
# ---------------------------------------------------------------------------
def _isb_candidate_urls(entry: dict) -> list[tuple[str, str, str]]:
    """Return [(url, expected_kind, variant_label)] in fallback order.
    For Acts with prefer_revised=true, try the LRC revised PDF first; then
    the original irishstatutebook.ie PDF; then the print HTML."""
    kind = entry["kind"]
    year = int(entry["year"])
    number = int(entry["number"])
    n_padded = f"{number:04d}"
    isb_pdf = f"{ISB_BASE}/pdf/{year}/en.{kind}.{year}.{n_padded}.pdf"
    enacted_or_made = "enacted" if kind == "act" else "made"
    isb_html = f"{ISB_BASE}/eli/{year}/{kind}/{number}/{enacted_or_made}/en/print.html"
    candidates: list[tuple[str, str, str]] = []
    if kind == "act" and entry.get("prefer_revised"):
        lrc_pdf = f"{LRC_REVISED_BASE}/eli/{year}/act/{number}/revised/en/pdf"
        candidates.append((lrc_pdf, "pdf", "revised"))
    candidates.append((isb_pdf, "pdf", "original"))
    candidates.append((isb_html, "html", "original"))
    return candidates


def download_irishstatutebook(entries: list[dict], index: dict[str, dict]) -> tuple[int, int, int]:
    """Download Irish Acts and Statutory Instruments. Source: irishstatutebook.ie
    plus revised PDFs from revisedacts.lawreform.ie for consolidated Acts."""
    DOCS_IE_PRIMARY.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    _isb_logger.info("=== Irish Statute Book crawl started ===")
    console.rule("[bold]Irish Statute Book[/bold]")

    with httpx.Client(
        headers={"User-Agent": USER_AGENT}, follow_redirects=True, timeout=120.0
    ) as client:
        for entry in entries:
            doc_id = entry["id"]
            kind = entry["kind"]
            short = entry["short_name"]
            title = entry["title"]
            year = entry["year"]
            doc_type = "act" if kind == "act" else "statutory_instrument"

            # Try candidate URLs in fallback order via HEAD probe
            chosen_url = chosen_ext = chosen_variant = None
            for url, expected_kind, variant in _isb_candidate_urls(entry):
                try:
                    _throttle(httpx.URL(url).host)
                    r = client.head(url, timeout=30.0)
                    if r.status_code != 200:
                        continue
                    ct = r.headers.get("content-type", "").lower()
                    if expected_kind == "pdf" and "pdf" not in ct:
                        continue
                    if expected_kind == "html" and "html" not in ct:
                        continue
                    chosen_url, chosen_ext, chosen_variant = url, expected_kind, variant
                    break
                except Exception as ex:
                    _isb_logger.warning(f"probe error {url}: {ex}")
                    continue

            now_iso = datetime.now(timezone.utc).isoformat()

            if not chosen_url:
                _isb_logger.error(f"FAILED {doc_id}: no candidate URL responded 200")
                console.print(f"[red]FAILED[/red] {doc_id}: no working URL")
                row = {
                    "id": doc_id, "source": "irishstatutebook", "short_name": short,
                    "title": title, "doc_type": doc_type,
                    "url": "", "published_date": str(year), "local_path": "",
                    "sha256": "", "downloaded_at": now_iso, "status": "failed",
                }
                index[doc_id] = row
                upsert_index_row(row)
                failed += 1
                continue

            local_path = DOCS_IE_PRIMARY / f"{doc_id}_{short}.{chosen_ext}"

            # Idempotency: existing file with matching index sha256 → unchanged
            if local_path.exists() and doc_id in index:
                existing_sha = sha256_of(local_path)
                if index[doc_id].get("sha256") == existing_sha:
                    _isb_logger.info(f"unchanged: {doc_id}")
                    index[doc_id] = dict(index[doc_id], status="unchanged")
                    upsert_index_row(index[doc_id])
                    unchanged += 1
                    continue

            try:
                r = _get(client, chosen_url)
                r.raise_for_status()
                local_path.write_bytes(r.content)
                sha = sha256_of(local_path)
                row = {
                    "id": doc_id,
                    "source": "irishstatutebook",
                    "short_name": short,
                    "title": title,
                    "doc_type": doc_type,
                    "url": chosen_url,
                    "published_date": str(year),
                    "local_path": str(local_path.relative_to(BASE_DIR)),
                    "sha256": sha,
                    "downloaded_at": now_iso,
                    "status": "ok",
                }
                index[doc_id] = row
                upsert_index_row(row)
                _isb_logger.info(
                    f"ok: {doc_id} ({chosen_variant} {chosen_ext}) {chosen_url} ({len(r.content)} bytes)"
                )
                console.print(
                    f"[green]ok[/green] {doc_id}  [dim]{chosen_variant} {chosen_ext} ({len(r.content):,} bytes)[/dim]"
                )
                new_count += 1
            except Exception as ex:
                _isb_logger.error(f"FAILED {doc_id}: {ex}")
                console.print(f"[red]FAILED[/red] {doc_id}: {ex}")
                row = {
                    "id": doc_id, "source": "irishstatutebook", "short_name": short,
                    "title": title, "doc_type": doc_type,
                    "url": chosen_url, "published_date": str(year), "local_path": "",
                    "sha256": "", "downloaded_at": now_iso, "status": "failed",
                }
                index[doc_id] = row
                upsert_index_row(row)
                failed += 1

    _isb_logger.info(
        f"=== Irish Statute Book crawl finished — new={new_count} unchanged={unchanged} failed={failed} ==="
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# Central Bank of Ireland crawler
# ---------------------------------------------------------------------------
def _cbi_slug_from_pdf_url(url: str) -> str:
    """URL → short slug derived from the PDF's filename basename."""
    path = urlparse(url).path
    fname = path.rsplit("/", 1)[-1]
    if fname.lower().endswith(".pdf"):
        fname = fname[:-4]
    slug = re.sub(r"[^a-z0-9]+", "_", fname.lower()).strip("_")[:60]
    return slug or "doc"


def _cbi_collect_pdf_links(client: httpx.Client, page_url: str) -> list[dict]:
    """Fetch a CBI index page and return [{url, title}] for every PDF link found."""
    try:
        r = _get(client, page_url)
        r.raise_for_status()
    except Exception as ex:
        _cbi_logger.warning(f"index page failed {page_url}: {ex}")
        return []
    soup = BeautifulSoup(r.text, "lxml")
    out: list[dict] = []
    seen_local: set[str] = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if ".pdf" not in href.lower():
            continue
        abs_url = urljoin(page_url, href).split("#")[0]
        if "centralbank.ie" not in urlparse(abs_url).netloc:
            continue
        if abs_url in seen_local:
            continue
        seen_local.add(abs_url)
        text = a.get_text(strip=True)[:200] or ""
        out.append({"url": abs_url, "title": text})
    return out


def _cbi_collect_hub_with_subpages(
    client: httpx.Client, hub_url: str, path_prefix: str
) -> list[dict]:
    """Walk a hub one level deep within `path_prefix`. Returns deduped PDF links
    from the hub itself and from every in-section subpage."""
    docs: list[dict] = []
    seen_urls: set[str] = set()

    # Pull the hub HTML once so we can extract both PDFs and subpage links.
    try:
        r = _get(client, hub_url)
        r.raise_for_status()
    except Exception as ex:
        _cbi_logger.warning(f"hub failed {hub_url}: {ex}")
        return docs

    soup = BeautifulSoup(r.text, "lxml")
    subpages: list[str] = []
    seen_subs: set[str] = set()
    for a in soup.find_all("a", href=True):
        abs_url = urljoin(hub_url, a["href"]).split("#")[0]
        p = urlparse(abs_url)
        if "centralbank.ie" not in p.netloc:
            continue
        if ".pdf" in abs_url.lower():
            if abs_url not in seen_urls:
                seen_urls.add(abs_url)
                docs.append({"url": abs_url, "title": a.get_text(strip=True)[:200]})
            continue
        if not p.path.startswith(path_prefix):
            continue
        if abs_url == hub_url:
            continue
        if abs_url in seen_subs:
            continue
        seen_subs.add(abs_url)
        subpages.append(abs_url)

    _cbi_logger.info(
        f"hub {hub_url} -> {len(docs)} top-level PDFs, {len(subpages)} subpages"
    )

    for sub_url in subpages:
        sub_pdfs = _cbi_collect_pdf_links(client, sub_url)
        added = 0
        for d in sub_pdfs:
            if d["url"] not in seen_urls:
                seen_urls.add(d["url"])
                docs.append(d)
                added += 1
        _cbi_logger.info(f"  subpage {sub_url} -> {len(sub_pdfs)} PDFs ({added} new)")

    return docs


def download_cbi(cbi_cfg: dict, index: dict[str, dict]) -> tuple[int, int, int]:
    """Crawl CBI index pages, dedupe PDF links, download each PDF."""
    DOCS_IE_CBI.mkdir(parents=True, exist_ok=True)
    new_count = unchanged = failed = 0

    _cbi_logger.info("=== CBI crawl started ===")
    console.rule("[bold]Central Bank of Ireland[/bold]")

    with httpx.Client(
        headers={"User-Agent": USER_AGENT}, follow_redirects=True, timeout=120.0
    ) as client:
        # Phase 1: collect unique PDF URLs from all index pages + direct list
        seen_urls: set[str] = set()
        all_docs: list[dict] = []

        index_pages = cbi_cfg.get("index_pages") or []
        for page_url in index_pages:
            links = _cbi_collect_pdf_links(client, page_url)
            _cbi_logger.info(f"index page {page_url} -> {len(links)} PDFs")
            console.print(f"[dim]index[/dim] {page_url} -> {len(links)} PDFs")
            for d in links:
                if d["url"] not in seen_urls:
                    seen_urls.add(d["url"])
                    all_docs.append(d)

        for direct_url in (cbi_cfg.get("direct_pdfs") or []):
            if direct_url not in seen_urls:
                seen_urls.add(direct_url)
                all_docs.append({"url": direct_url, "title": ""})

        for hub_entry in (cbi_cfg.get("subpage_hubs") or []):
            hub_url = hub_entry["hub"]
            prefix = hub_entry["path_prefix"]
            sub_docs = _cbi_collect_hub_with_subpages(client, hub_url, prefix)
            new_in_hub = 0
            for d in sub_docs:
                if d["url"] not in seen_urls:
                    seen_urls.add(d["url"])
                    all_docs.append(d)
                    new_in_hub += 1
            _cbi_logger.info(
                f"hub total {hub_url}: {len(sub_docs)} PDFs found ({new_in_hub} new vs prior pages)"
            )
            console.print(
                f"[dim]hub[/dim] {hub_url} -> {len(sub_docs)} PDFs ({new_in_hub} new)"
            )

        _cbi_logger.info(f"total unique PDFs to attempt: {len(all_docs)}")
        console.print(f"[bold]CBI: {len(all_docs)} unique PDFs to fetch[/bold]")

        # Phase 2: rebuild slug -> url claims from the index. Idempotency
        # requires that the same URL re-claims its prior slug on re-run, so we
        # track the URL each slug was claimed for, not just the bare slug.
        slug_to_url: dict[str, str] = {}
        for r_idx in index.values():
            if r_idx.get("source") == "cbi":
                s = r_idx.get("short_name", "")
                u = r_idx.get("url", "")
                if s:
                    slug_to_url[s] = u

        # Phase 3: per-doc download
        for d in all_docs:
            url = d["url"]
            base_slug = _cbi_slug_from_pdf_url(url)
            slug = base_slug
            counter = 1
            # Only bump suffix when the slug is claimed for a *different* URL
            while slug in slug_to_url and slug_to_url[slug] != url:
                slug = f"{base_slug}_{counter}"
                counter += 1
            slug_to_url[slug] = url
            doc_id = f"cbi:{slug}"
            local_path = DOCS_IE_CBI / f"cbi_{slug}.pdf"
            now_iso = datetime.now(timezone.utc).isoformat()

            # Idempotency
            if local_path.exists() and doc_id in index:
                existing_sha = sha256_of(local_path)
                if index[doc_id].get("sha256") == existing_sha:
                    _cbi_logger.info(f"unchanged: {slug}")
                    index[doc_id] = dict(index[doc_id], status="unchanged")
                    upsert_index_row(index[doc_id])
                    unchanged += 1
                    continue

            try:
                r = _get(client, url)
                r.raise_for_status()
                content = r.content
                if not content.startswith(b"%PDF"):
                    raise ValueError(f"Not a PDF (first bytes: {content[:20]!r})")
                local_path.write_bytes(content)
                sha = sha256_of(local_path)
                row = {
                    "id": doc_id,
                    "source": "cbi",
                    "short_name": slug,
                    "title": d.get("title") or slug.replace("_", " "),
                    "doc_type": "guidance",
                    "url": url,
                    "published_date": "",
                    "local_path": str(local_path.relative_to(BASE_DIR)),
                    "sha256": sha,
                    "downloaded_at": now_iso,
                    "status": "ok",
                }
                index[doc_id] = row
                upsert_index_row(row)
                _cbi_logger.info(f"ok: cbi_{slug}.pdf  {url}  ({len(content)} bytes)")
                console.print(f"[green]ok[/green] cbi_{slug}.pdf  [dim]({len(content):,} bytes)[/dim]")
                new_count += 1
            except Exception as ex:
                _cbi_logger.error(f"FAILED cbi_{slug}: {ex}  url={url}")
                console.print(f"[red]FAILED[/red] cbi_{slug}: {ex}")
                row = {
                    "id": doc_id, "source": "cbi", "short_name": slug,
                    "title": d.get("title", ""), "doc_type": "guidance",
                    "url": url, "published_date": "", "local_path": "",
                    "sha256": "", "downloaded_at": now_iso, "status": "failed",
                }
                index[doc_id] = row
                upsert_index_row(row)
                failed += 1

    _cbi_logger.info(
        f"=== CBI crawl finished — new={new_count} unchanged={unchanged} failed={failed} ==="
    )
    return new_count, unchanged, failed


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------
def run(source_filter: str | None) -> None:
    with SOURCES_YAML.open(encoding="utf-8") as fh:
        config = yaml.safe_load(fh)

    index = load_index()
    total_new = total_unchanged = total_failed = 0

    sources_to_run = source_filter or "all"
    console.rule(f"[bold]EU Regulatory Corpus[/bold] — source={sources_to_run}")

    if sources_to_run in ("all", "eurlex"):
        eurlex_entries = config.get("eurlex", [])
        n, u, f = download_eurlex(eurlex_entries, index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "eba"):
        n, u, f = download_eba(index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "ecb"):
        n, u, f = download_ecb(index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "esma"):
        n, u, f = download_esma(index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "fatf"):
        n, u, f = download_fatf(index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "ireland", "irishstatutebook"):
        ireland_primary = (
            config.get("national", {}).get("ireland", {}).get("primary", [])
        )
        n, u, f = download_irishstatutebook(ireland_primary, index)
        total_new += n
        total_unchanged += u
        total_failed += f

    if sources_to_run in ("all", "ireland", "cbi"):
        cbi_cfg = (
            config.get("national", {}).get("ireland", {}).get("cbi", {})
        )
        n, u, f = download_cbi(cbi_cfg, index)
        total_new += n
        total_unchanged += u
        total_failed += f

    save_index(index)

    console.rule()
    console.print(
        f"[bold]Summary:[/bold] "
        f"[green]{total_new} new[/green], "
        f"[dim]{total_unchanged} unchanged[/dim], "
        f"[red]{total_failed} failed[/red]"
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Download EU regulatory documents.")
    parser.add_argument(
        "--source",
        choices=["eurlex", "eba", "ecb", "esma", "fatf", "irishstatutebook", "cbi", "ireland"],
        default=None,
        help="Download only this source (default: all)",
    )
    args = parser.parse_args()
    run(args.source)
